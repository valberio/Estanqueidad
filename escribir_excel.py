import argparse
import os
import sys
import json
import openpyxl

def abrir_planilla(planilla_path):
    """
    Abre la planilla de Excel y devuelve un objeto workbook manipulable.
    """
    try:
        workbook = openpyxl.load_workbook(planilla_path)
        return workbook
    except Exception as e:
        print(f"Error abriendo la planilla: {e}")
        return None

def encontrar_columna(planilla, col_titulo):
    """
    Encontrar el número de índice de la columna con un cierto título en la planilla.
    """
    for row in planilla.iter_rows(min_row=7, max_row=7): #Itero en la fila 7 porque ahi estan los nombres de los sensores
        for cell in row:
            if cell.value == col_titulo:
                return cell.column

    print(f"Columna '{col_titulo}' no encontrada.")
    return None

def escribir_en_columna(planilla, col_titulo, datos):
    """
    Escribe la lista de datos en la columna con el título pasado por parámetro.
    """
    indice_columna = encontrar_columna(planilla, col_titulo)

    if indice_columna is not None:
        for row, value in enumerate(datos, start=9):
            print(f"Escribi {value}")
            planilla.cell(row=row, column=indice_columna, value=value)

def guardar_planilla(workbook, planilla_path):
    """
    Guardar el workbook en el path. La planilla y el workbook son lo mismo.
    """
    try:
        planilla_full_path = os.path.abspath(planilla_path)
        workbook.save(planilla_path)
        print("Planilla guardada.")
    except Exception as e:
        print(f"Error guardando la planilla: {e}")

def fila_numero_de_medicion(workbook, numero_medicion):
    """
    Busca en la columna de números de medición un número de medición específico; y devuelve la fila en la que se encuentra. 
    """
    for cell in workbook['A']:
        print(f"{cell.value}")
        if str(cell.value) == str(numero_medicion):
            print(f"{cell.value}")
            return cell.row
    print("El numero de medición no se encontró")
    return None

def escribir_medicion_especifica(workbook, titulo_columna, numero_medicion, medicion):
    """
    Escribe el valor de un número de medición en particular de un sensor. El número de medición debe estar previamente cargado en el Excel
    """
    fila = fila_numero_de_medicion(workbook, numero_medicion)
    columna = encontrar_columna(workbook, titulo_columna)

    if fila is not None and columna is not None:
        workbook.cell(row=fila, column=columna, value=medicion)
    else:
        print("Fallo la carga de la medicion.")
    

def main():
    planilla_path = 'Reporte_CNE 01_01_2019_00_11_51 (1) (1).xlsx'

    nombre_hoja = 'Tabla CNEA'

    workbook = abrir_planilla(planilla_path)
    planilla = workbook[nombre_hoja]

    if len(sys.argv) < 2:
        print("---MODO DE USO---")
        print("- Modo para escribir la columna de mediciones COMPLETA de un sensor:")
        print("  python escribir_excel.py columna nombre_sensor [medicion1, medicion2, ...]")
        print("- Escribir una MEDICION ESPECÍFICA de un sensor:")
        print("  python escribir_excel.py medicion_particular nombre_sensor numero_medicion medicion")
        return

    if workbook is None:
        print(f"Error abriendo la planilla de Excel")
        return None
        
    
    if sys.argv[1] == "columna":
        titulo_columna = sys.argv[2]
        datos = json.loads(sys.argv[3])
        print(datos)
        escribir_en_columna(planilla, titulo_columna, datos)
        guardar_planilla(workbook, planilla_path)

    elif sys.argv[1] == "medicion_particular":
        print("Medicion particular")
        titulo_columna = sys.argv[2]
        numero_medicion = sys.argv[3]
        medicion = sys.argv[4]
        escribir_medicion_especifica(planilla, titulo_columna, numero_medicion, medicion)
        guardar_planilla(workbook, planilla_path)

    else:
        print("---MODO DE USO---")
        print("- Modo para escribir la columna de mediciones COMPLETA de un sensor:")
        print("  python escribir_excel.py columna nombre_sensor [medicion1, medicion2, ...]")
        print("- Escribir una MEDICION ESPECÍFICA de un sensor:")
        print("  python escribir_excel.py medicion_particular nombre_sensor numero_medicion medicion")

if __name__ == "__main__":
    main()